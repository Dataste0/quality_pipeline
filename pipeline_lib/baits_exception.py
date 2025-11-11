import os
import pandas as pd
import re
import sys
from openpyxl.styles import NumberFormatDescriptor
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import pipeline_lib.config as cfg
import pipeline_lib.pipeline_utils as pu

RAWDATA_BASE_FOLDER = cfg.RAWDATA_ROOT_PATH
OLAP_BASE_FOLDER = cfg.OLAP_EXPORT_DIR_PATH

# --- Logger
import logging
logger = logging.getLogger(__name__)

# --- Setup Project List
PROJECT_MASTERFILE = cfg.PROJECT_INFO_FILE_PATH
project_list_df = pu.load_project_info(PROJECT_MASTERFILE, active_only=False)



CBV2_PROJECT_ID = "a01Hs00001ocUa0IAE"
TARGET_CBV2 = 0.64

EB_PROJECT_ID = "a01Hs00001ocUZgIAM"
TARGET_EB = 0.85

# Compile file list patterns
def file_list_compile(project_id, weekdate_str):
    filedate = datetime.strptime(weekdate_str, '%Y-%m-%d') + timedelta(days=3)
    filedate = filedate.strftime('%Y-%m-%d')
    if project_id == CBV2_PROJECT_ID:
        files_required = [
            f"cvs_cb_withhold_metrics_overview_{filedate}.csv",
            f"cvs_cb_withhold_rater_level_metrics_{filedate}.csv",
            f"cvs_cb_exagg_metrics_overview_{filedate}.csv",
            f"cvs_cb_exagg_rater_level_metrics_{filedate}.csv"
        ]
    elif project_id == EB_PROJECT_ID:
        files_required = [
            f"cvs_engagement_bait_metrics_overview_{filedate}.csv",
            f"cvs_engagement_bait_rater_level_metrics_{filedate}.csv"
        ]
    else:
        raise ValueError(f"Unknown project: {project_id}")
    return files_required


# Check if files are ready
def all_files_ready(files_required: list, filepath: str) -> bool:
    found = 0
    #print(f"---Filerequired: {files_required}\n")
    for file_required in files_required:
        path = os.path.join(filepath, file_required)
        #print(f"---Path: {path}\n")
        if os.path.exists(path):
            #print(f"Found required file: {path}")
            found += 1
    return found == len(files_required)


### GENERATE REPORTS CBV2 ###
def generate_reports_CB(weekdatestr: str, path: str):
        
    def prepare_overview_df(df, lb):
        df.columns = df.columns.str.strip()
        last_date = df['sample_ds'].max()
        df_last_date = df[df['sample_ds'] == last_date].copy()
        df_last_date['main_label'] = lb

        df_pos = df_last_date[df_last_date['label_category_name'].str.contains('positive', case=False)]
        df = df_pos.copy()
        df['lang'] = df['routing_name'].str.split('_').str.get(2)
        df['lang'] = df['lang'].str.upper()
        selected_columns = ['sample_ds', 'routing_name', 'lang', 'main_label', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']
        df = df[selected_columns]
        return df

    def prepare_raters_df(df, lb):
        df.columns = df.columns.str.strip()
        last_date = df['sample_ds'].max()
        df_last_date = df[df['sample_ds'] == last_date].copy()
        df_last_date['main_label'] = lb
        df = df_last_date
        df['lang'] = df['routing_name'].str.split('_').str.get(2)
        df['lang'] = df['lang'].str.upper()
        df['audit_count'] = df['tp_count'] + df['fp_count'] + df['fn_count']
        
        selected_columns = ['sample_ds', 'rater_id', 'lang', 'main_label', 'total_count', 'audit_count', 'tp_count', 'fp_count', 'tn_count', 'fn_count', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']
        df = df[selected_columns]
        return df


    #-- Function body --
    filedate = datetime.strptime(weekdatestr, '%Y-%m-%d') + timedelta(days=3)
    filedate = filedate.strftime('%Y-%m-%d')

    # overview consolidated
    overview_dfs = []
    for lb in ['withhold', 'exagg']:
        report_file_overview = os.path.join(path, f"cvs_cb_{lb}_metrics_overview_{filedate}.csv")
        df = pd.read_csv(report_file_overview)
        df = prepare_overview_df(df, lb)
        overview_dfs.append(df)
    overview_cons_df = pd.concat(overview_dfs)

    # raters consolidated
    raters_dfs = []
    for lb in ['withhold', 'exagg']:
        report_file_raters = os.path.join(path, f"cvs_cb_{lb}_rater_level_metrics_{filedate}.csv")
        df = pd.read_csv(report_file_raters)
        df = prepare_raters_df(df, lb)
        raters_dfs.append(df)
    raters_cons_df = pd.concat(raters_dfs)

    # Build - Combined Report
    combined_tmp_df = overview_cons_df[overview_cons_df['routing_name'].str.contains('combined', case=False)].copy()
    combined_tmp_df.drop(columns=['lang'], inplace=True)
    combined_tmp_df.rename(columns= {'routing_name': 'lang'}, inplace=True)
    combined_df = combined_tmp_df.groupby(['sample_ds', 'lang'])[['unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']].mean().reset_index()
    combined_df.insert(2, 'main_label', 'withhold+exagg')
    #combined_df.head()

    # Build - Project-Label Report
    project_label_df = combined_tmp_df.copy()
    #project_label_df.head()

    # Build - Rater-Label Report
    raters_cons_df['above_target'] = raters_cons_df['unweighted_f1']>=TARGET_CBV2
    raters_cons_df['has_valid_score'] = raters_cons_df['tp_count'] + raters_cons_df['fp_count'] + raters_cons_df['fn_count'] > 0
    raters_cons_df['positivity'] = (raters_cons_df['tp_count'] + raters_cons_df['fp_count']) / raters_cons_df['total_count']
    raters_cons_df.head(100)
    rater_label_df = raters_cons_df.copy()
    rater_label_df['rater_id'] = "'" + rater_label_df['rater_id'].astype(str)
    #rater_label_df.head(100)

    # Build Rater Report
    selected_cols = ['sample_ds', 'rater_id', 'lang', 'main_label', 'total_count', 'audit_count', 'tp_count', 'fp_count', 'tn_count', 'fn_count', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall', 'positivity']
    #print(f"rater_label_df columns: {rater_label_df.columns.tolist()}")
    rater_tmp_df = rater_label_df[selected_cols]
    rater_df = rater_tmp_df.groupby(['sample_ds', 'rater_id', 'lang']).agg({
        'total_count': 'sum',
        'audit_count': 'sum',
        'tp_count': 'sum',
        'fp_count': 'sum',
        'tn_count': 'sum',
        'fn_count': 'sum',
        'unweighted_accuracy': 'mean',
        'unweighted_f1': 'mean',
        'unweighted_precision': 'mean',
        'unweighted_recall': 'mean',
        'positivity': 'mean'
    }).reset_index()
    rater_df['above_target'] = rater_df['unweighted_f1']>=TARGET_CBV2
    rater_df['has_valid_score'] = rater_df['tp_count'] + rater_df['fp_count'] + rater_df['fn_count'] > 0
    rater_df.insert(3, 'main_label', 'withhold+exagg')
    #rater_df['rater_id'] = "'" + rater_df['rater_id'].astype(str)
    #rater_df.head(100)

    # Aggregated from Rater Label Report
    rater_label_tmp = rater_label_df[rater_label_df['has_valid_score'].astype(bool)].copy()
    rater_label_tmp['above_target_num'] = rater_label_tmp['above_target'].astype(int)

    rater_agg_df = rater_label_tmp.groupby(['sample_ds', 'lang', 'main_label']).agg({
        'rater_id': 'count',
        'above_target_num': 'sum',
        'total_count': 'sum',
        'audit_count': 'sum',
        'tp_count': 'sum',
        'fp_count': 'sum',
        'tn_count': 'sum',
        'fn_count': 'sum',
        'unweighted_accuracy': 'mean',
        'unweighted_f1': 'mean',
        'unweighted_precision': 'mean',
        'unweighted_recall': 'mean',
        'positivity' : 'mean'
    }).reset_index()
    rater_agg_df.rename(columns={'rater_id' : 'rater_count', 'above_target_num': 'raters_above_target'}, inplace=True)
    #rater_agg_df.head(100)


    # Aggregated from Rater Combined Report
    rater_tmp = rater_df[rater_df['has_valid_score'].astype(bool)].copy()
    rater_tmp['above_target_num'] = rater_tmp['above_target'].astype(int)

    rater_agg_comb_df = rater_tmp.groupby(['sample_ds', 'lang']).agg({
        'rater_id': 'count',
        'above_target_num': 'sum',
        'total_count': 'sum',
        'audit_count': 'sum',
        'tp_count': 'sum',
        'fp_count': 'sum',
        'tn_count': 'sum',
        'fn_count': 'sum',
        'unweighted_accuracy': 'mean',
        'unweighted_f1': 'mean',
        'unweighted_precision': 'mean',
        'unweighted_recall': 'mean',
        'positivity': 'mean'
    }).reset_index()
    rater_agg_comb_df.rename(columns={'rater_id' : 'rater_count', 'above_target_num': 'raters_above_target'}, inplace=True)
    rater_agg_comb_df.insert(2, 'main_label', 'withhold+exagg')
    rater_agg_comb_df['raters_above_target_perc'] = rater_agg_comb_df['raters_above_target'] / rater_agg_comb_df['rater_count']
    #rater_agg_comb_df.head(100)


    # Build - Market-Label Report
    market_label_df = overview_cons_df[~overview_cons_df['routing_name'].str.contains('combined', case=False, na=False)].copy()
    market_label_df.drop(columns=['routing_name'], inplace=True)
    #market_label_df.head(200)


    # Build - Market-Combined Report
    market_df = market_label_df.copy()
    market_df = market_df.groupby(['sample_ds', 'lang']).agg({
        'unweighted_accuracy': 'mean',
        'unweighted_f1': 'mean',
        'unweighted_precision': 'mean',
        'unweighted_recall': 'mean'
    }).reset_index()
    market_df.insert(2, 'main_label', 'withhold+exagg')


    # market_df left join rater_agg_comb_df
    market_df = pd.merge(market_df, rater_agg_comb_df[['sample_ds', 'lang', 'rater_count', 'raters_above_target', 'raters_above_target_perc', 'audit_count', 'positivity']], on=['sample_ds', 'lang'], how='left')
    #market_df.head(100)

    # Join audit count to Overall Combined
    rater_agg_comb_overall_df = rater_agg_comb_df.groupby(['sample_ds', 'main_label']).agg({
        'rater_count': 'sum',
        'raters_above_target': 'sum',
        'audit_count': 'sum'
    }).reset_index()

    combined_df = pd.merge(combined_df, rater_agg_comb_overall_df[['sample_ds', 'main_label', 'audit_count', 'rater_count', 'raters_above_target']], on=['sample_ds', 'main_label'], how='left')
    combined_df['raters_above_target_perc'] = combined_df['raters_above_target'] / combined_df['rater_count']
    combined_df.head(100)
    #market_df = pd.merge(market_df, rater_agg_comb_df[['sample_ds', 'lang', 'rater_count', 'raters_above_target', 'raters_above_target_perc', 'audit_count']], on=['sample_ds', 'lang'], how='left')


    # Excel Export
    xlsx_file_name = f"CB Consolidated Report WE{weekdatestr}.xlsx"
    xlsx_file_path = os.path.join(path, xlsx_file_name)
    #print(f"Creating Excel file: {xlsx_file_path}")

    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        # Saving reports
        project_label_df.to_excel(writer, sheet_name='Overall-Label', index=False)
        combined_df.to_excel(writer, sheet_name='Overall-Combined', index=False)

        rater_label_df.to_excel(writer, sheet_name='Rater-Label', index=False)
        rater_df.to_excel(writer, sheet_name='Rater-Combined', index=False)

        market_label_df.to_excel(writer, sheet_name='Market-Label', index=False)
        market_df.to_excel(writer, sheet_name='Market-Combined', index=False)

        # Changes col format
        pct_cols = ['unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall', 'raters_above_target_perc', 'positivity']
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            df = None

            if sheet_name == 'Overall-Label':
                df = project_label_df
            elif sheet_name == 'Overall-Combined':
                df = combined_df
            elif sheet_name == 'Rater-Label':
                df = rater_label_df
            elif sheet_name == 'Rater-Combined':
                df = rater_df
            elif sheet_name == 'Market-Label':
                df = market_label_df
            elif sheet_name == 'Market-Combined':
                df = market_df

            if df is not None:
                for col_name in pct_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        # Starts from 2nd row
                        for row in range(2, len(df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = '0.00%'
            else:
                print(f"ERROR: Sheet '{sheet_name}' not found in the Excel file.")
        #print("Excel file created successfully.")


    # Dashboard report smr-workflow
    dash_workflow_df = market_df.copy()

    #date_obj = datetime.strptime(weekdatestr, '%Y-%m-%d')
    #we_date_obj = date_obj - timedelta(days=3)
    #we_date = pd.to_datetime(we_date_obj).strftime("%Y-%m-%d")

    dash_workflow_df.insert(0, 'week_ending', weekdatestr)
    dash_workflow_df.insert(1, 'project_id', CBV2_PROJECT_ID)

    dash_workflow_df['target_goal'] = TARGET_CBV2
    dash_workflow_df['raters_above_target_f1'] = dash_workflow_df['raters_above_target']
    dash_workflow_df['audited_instances'] = dash_workflow_df['audit_count']
    dash_workflow_df['audited_instances_f1'] = dash_workflow_df['audit_count']

    dash_workflow_df.rename(columns={
            'lang': 'workflow',
            'unweighted_accuracy': 'workflow_score',
            'unweighted_f1': 'workflow_f1score',
            'unweighted_precision': 'workflow_precision',
            'unweighted_recall': 'workflow_recall',
        }, inplace=True)
    dash_workflow_df.drop(columns=['raters_above_target_perc', 'audit_count'], inplace=True)
    dash_workflow_df = pd.merge(dash_workflow_df, combined_df[['sample_ds', 'main_label', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']], on=['sample_ds', 'main_label'], how='left')

    dash_workflow_df.rename(columns={
            'unweighted_accuracy': 'project_score',
            'unweighted_f1': 'project_f1score',
            'unweighted_precision': 'project_precision',
            'unweighted_recall': 'project_recall',
        }, inplace=True)

    # remove samples_ds, main_label
    dash_workflow_df.drop(columns=['sample_ds', 'main_label', 'positivity'], inplace=True)

    dash_workflow_df['auditor_count'] = pd.NA
    dash_workflow_df['job_instances'] = pd.NA

    output_filepath = os.path.join(OLAP_BASE_FOLDER, CBV2_PROJECT_ID, weekdatestr, f"{CBV2_PROJECT_ID}_{weekdatestr}_audit_smr-workflow.csv")
    dash_workflow_df.to_csv(output_filepath, index=False, encoding='utf-8')


    # Dashboard report smr-rater-label
    dash_rater_df = rater_label_df.copy()

    #we_date_obj = datetime.strptime(datestr, '%Y-%m-%d') - timedelta(days=3)
    #we_date = pd.to_datetime(we_date_obj).strftime("%Y-%m-%d")

    dash_rater_df.insert(0, 'week_ending', weekdatestr)
    dash_rater_df.insert(1, 'project_id', CBV2_PROJECT_ID)

    dash_rater_df['parent_label'] = dash_rater_df['main_label']
    dash_rater_df.rename(columns={
            'lang': 'workflow',
            'total_count': 'tot_labels',
            'unweighted_accuracy': 'rater_label_score',
            'unweighted_f1': 'rater_label_f1score',
            'unweighted_precision': 'rater_label_precision',
            'unweighted_recall': 'rater_label_recall',
        }, inplace=True)
    dash_rater_df['correct_labels'] = dash_rater_df['tp_count'] + dash_rater_df['tn_count']

    dash_rater_df.drop(columns=['above_target', 'has_valid_score', 'positivity'], inplace=True)

    dash_rater_df = pd.merge(dash_rater_df, rater_df[['sample_ds', 'rater_id', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']], on=['sample_ds', 'rater_id'], how='left')

    dash_rater_df.rename(columns={
            'unweighted_accuracy': 'rater_score',
            'unweighted_f1': 'rater_f1score',
            'unweighted_precision': 'rater_precision',
            'unweighted_recall': 'rater_recall',
        }, inplace=True)

    dash_rater_df['rater_id'] = dash_rater_df['rater_id'].str.strip("'")

    dash_rater_df.drop(columns=['sample_ds', 'main_label'], inplace=True)

    output_filepath = os.path.join(OLAP_BASE_FOLDER, CBV2_PROJECT_ID, weekdatestr, f"{CBV2_PROJECT_ID}_{weekdatestr}_audit_smr-rater-label.csv")
    dash_rater_df.to_csv(output_filepath, index=False, encoding='utf-8')
    return True


### GENERATE REPORTS EB ###
def generate_reports_EB(weekdatestr: str, path: str):
    def prepare_overview_df(df):
        df.columns = df.columns.str.strip()
        last_date = df['sample_ds'].max()
        df_last_date = df[df['sample_ds'] == last_date].copy()
        df_last_date['main_label'] = 'eb'

        df_pos = df_last_date[df_last_date['label_category_name'].str.contains('positive', case=False)]
        df = df_pos.copy()
        df['lang'] = df['routing_name'].str.split('_').str.get(1)
        #df['lang'] = df['lang'].str.upper()
        selected_columns = ['sample_ds', 'routing_name', 'lang', 'main_label', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']
        df = df[selected_columns]
        return df

    def prepare_raters_df(df):
        df.columns = df.columns.str.strip()
        last_date = df['sample_ds'].max()
        df_last_date = df[df['sample_ds'] == last_date].copy()
        df_last_date['main_label'] = 'eb'
        df = df_last_date
        df['lang'] = df['routing_name'].str.split('_').str.get(1)
        #df['lang'] = df['lang'].str.upper()
        df['audit_count'] = df['tp_count'] + df['fp_count'] + df['fn_count']

        selected_columns = ['sample_ds', 'rater_id', 'lang', 'main_label', 'total_count', 'audit_count', 'tp_count', 'fp_count', 'tn_count', 'fn_count', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']
        df = df[selected_columns]
        return df
    
    #-- Function body --
    filedate = datetime.strptime(weekdatestr, '%Y-%m-%d') + timedelta(days=3)
    filedate = filedate.strftime('%Y-%m-%d')

    # overview
    report_file_overview = os.path.join(path, f"cvs_engagement_bait_metrics_overview_{filedate}.csv")
    overview_df = pd.read_csv(report_file_overview)
    overview_df = prepare_overview_df(overview_df)

    # raters
    report_file_raters = os.path.join(path, f"cvs_engagement_bait_rater_level_metrics_{filedate}.csv")
    raters_df = pd.read_csv(report_file_raters)
    raters_df = prepare_raters_df(raters_df)

    # Build - Overall Report
    overall_df = overview_df[overview_df['routing_name'].str.contains('combined', case=False)].copy()
    overall_df.drop(columns=['lang'], inplace=True)
    overall_df.rename(columns= {'routing_name': 'lang'}, inplace=True)
    
    # Build - Rater-Label Report
    raters_df['above_target'] = raters_df['unweighted_f1']>=TARGET_EB
    raters_df['has_valid_score'] = raters_df['tp_count'] + raters_df['fp_count'] + raters_df['fn_count'] > 0
    raters_df['positivity'] = (raters_df['tp_count'] + raters_df['fp_count']) / raters_df['total_count']
    raters_df = raters_df.copy()
    raters_df['rater_id'] = "'" + raters_df['rater_id'].astype(str)

    # Aggregated from Rater Combined Report
    rater_tmp = raters_df[raters_df['has_valid_score'].astype(bool)].copy()
    rater_tmp['above_target_num'] = rater_tmp['above_target'].astype(int)

    rater_agg_df = rater_tmp.groupby(['sample_ds', 'lang']).agg({
        'rater_id': 'count',
        'above_target_num': 'sum',
        'total_count': 'sum',
        'audit_count': 'sum',
        'tp_count': 'sum',
        'fp_count': 'sum',
        'tn_count': 'sum',
        'fn_count': 'sum',
        'unweighted_accuracy': 'mean',
        'unweighted_f1': 'mean',
        'unweighted_precision': 'mean',
        'unweighted_recall': 'mean',
        'positivity': 'mean'
    }).reset_index()
    rater_agg_df.rename(columns={'rater_id' : 'rater_count', 'above_target_num': 'raters_above_target'}, inplace=True)
    rater_agg_df.insert(2, 'main_label', 'eb')
    rater_agg_df['raters_above_target_perc'] = rater_agg_df['raters_above_target'] / rater_agg_df['rater_count']
    
    # Build - Market-Label Report
    market_df = overview_df[~overview_df['routing_name'].str.contains('combined', case=False, na=False)].copy()
    market_df.drop(columns=['routing_name'], inplace=True)

    # market_df left join rater_agg_comb_df
    market_df = pd.merge(market_df, rater_agg_df[['sample_ds', 'lang', 'rater_count', 'raters_above_target', 'raters_above_target_perc', 'audit_count', 'positivity']], on=['sample_ds', 'lang'], how='left')
    
    # Join audit count to Overall Combined
    rater_agg_overall_df = rater_agg_df.groupby(['sample_ds', 'main_label']).agg({
        'rater_count': 'sum',
        'raters_above_target': 'sum',
        'audit_count': 'sum',
        'positivity': 'mean'
    }).reset_index()

    overall_df = pd.merge(overall_df, rater_agg_overall_df[['sample_ds', 'main_label', 'audit_count', 'rater_count', 'raters_above_target', 'positivity']], on=['sample_ds', 'main_label'], how='left')
    overall_df['raters_above_target_perc'] = overall_df['raters_above_target'] / overall_df['rater_count']
    

    # Excel Export
    #we_date_obj = datetime.strptime(datestr, '%Y-%m-%d') - timedelta(days=3)
    #we_datestr = we_date_obj.strftime('WE%Y-%m-%d')
    xlsx_file_name = f"EB Consolidated Report WE{weekdatestr}.xlsx"
    xlsx_file_path = os.path.join(path, xlsx_file_name)
    print(f"Creating Excel file: {xlsx_file_path}")

    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        # Saving reports
        overall_df.to_excel(writer, sheet_name='Overall', index=False)
        raters_df.to_excel(writer, sheet_name='Raters', index=False)
        market_df.to_excel(writer, sheet_name='Market', index=False)

        # Changes col format
        pct_cols = ['unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall', 'raters_above_target_perc', 'positivity']
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            df = None

            if sheet_name == 'Overall':
                df = overall_df
            elif sheet_name == 'Raters':
                df = raters_df
            elif sheet_name == 'Market':
                df = market_df

            if df is not None:
                for col_name in pct_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        # Starts from 2nd row
                        for row in range(2, len(df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = '0.00%'
            else:
                print(f"ERROR: Sheet '{sheet_name}' not found in the Excel file.")
        #print("Excel file created successfully.")



    # Dashboard report smr-workflow
    dash_workflow_df = market_df.copy()

    #date_obj = datetime.strptime(weekdatestr, '%Y-%m-%d')
    #we_date_obj = date_obj - timedelta(days=3)
    #we_date = pd.to_datetime(we_date_obj).strftime("%Y-%m-%d")

    dash_workflow_df.insert(0, 'week_ending', weekdatestr)
    dash_workflow_df.insert(1, 'project_id', EB_PROJECT_ID)

    dash_workflow_df['target_goal'] = TARGET_EB
    dash_workflow_df['raters_above_target_f1'] = dash_workflow_df['raters_above_target']
    dash_workflow_df['audited_instances'] = dash_workflow_df['audit_count']
    dash_workflow_df['audited_instances_f1'] = dash_workflow_df['audit_count']

    dash_workflow_df.rename(columns={
            'lang': 'workflow',
            'unweighted_accuracy': 'workflow_score',
            'unweighted_f1': 'workflow_f1score',
            'unweighted_precision': 'workflow_precision',
            'unweighted_recall': 'workflow_recall',
        }, inplace=True)
    dash_workflow_df.drop(columns=['raters_above_target_perc', 'audit_count'], inplace=True)
    dash_workflow_df = pd.merge(dash_workflow_df, overall_df[['sample_ds', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']], on=['sample_ds'], how='left')

    dash_workflow_df.rename(columns={
            'unweighted_accuracy': 'project_score',
            'unweighted_f1': 'project_f1score',
            'unweighted_precision': 'project_precision',
            'unweighted_recall': 'project_recall',
        }, inplace=True)

    # remove samples_ds, main_label
    dash_workflow_df.drop(columns=['sample_ds', 'main_label', 'positivity'], inplace=True)

    dash_workflow_df['auditor_count'] = pd.NA
    dash_workflow_df['job_instances'] = pd.NA

    output_filepath = os.path.join(OLAP_BASE_FOLDER, EB_PROJECT_ID, weekdatestr, f"{EB_PROJECT_ID}_{weekdatestr}_audit_smr-rater-label.csv")
    dash_workflow_df.to_csv(output_filepath, index=False, encoding='utf-8')


    # Dashboard report smr-rater-label
    dash_rater_df = raters_df.copy()

    #date_obj = datetime.strptime(datestr, '%Y-%m-%d')
    #we_date_obj = date_obj - timedelta(days=3)
    #we_date = pd.to_datetime(we_date_obj).strftime("%Y-%m-%d")

    dash_rater_df.insert(0, 'week_ending', weekdatestr)
    dash_rater_df.insert(1, 'project_id', EB_PROJECT_ID)

    dash_rater_df['parent_label'] = dash_rater_df['main_label']
    dash_rater_df.rename(columns={
            'lang': 'workflow',
            'total_count': 'tot_labels',
            'unweighted_accuracy': 'rater_label_score',
            'unweighted_f1': 'rater_label_f1score',
            'unweighted_precision': 'rater_label_precision',
            'unweighted_recall': 'rater_label_recall',
        }, inplace=True)
    dash_rater_df['correct_labels'] = dash_rater_df['tp_count'] + dash_rater_df['tn_count']

    dash_rater_df.drop(columns=['above_target', 'has_valid_score', 'positivity'], inplace=True)

    dash_rater_df = pd.merge(dash_rater_df, raters_df[['sample_ds', 'rater_id', 'unweighted_accuracy', 'unweighted_f1', 'unweighted_precision', 'unweighted_recall']], on=['sample_ds', 'rater_id'], how='left')

    dash_rater_df.rename(columns={
            'unweighted_accuracy': 'rater_score',
            'unweighted_f1': 'rater_f1score',
            'unweighted_precision': 'rater_precision',
            'unweighted_recall': 'rater_recall',
        }, inplace=True)

    dash_rater_df['rater_id'] = dash_rater_df['rater_id'].str.strip("'")

    dash_rater_df.drop(columns=['sample_ds', 'main_label'], inplace=True)

    output_filepath = os.path.join(OLAP_BASE_FOLDER, EB_PROJECT_ID, weekdatestr, f"{EB_PROJECT_ID}_{weekdatestr}_audit_smr-rater-label.csv")
    dash_rater_df.to_csv(output_filepath, index=False, encoding='utf-8')
    dash_rater_df.to_csv('debug_eb_rater.csv', index=False, encoding='utf-8')
    return True










def overwrite_olap(project_id, reporting_week):
    reporting_week = pd.to_datetime(reporting_week, errors='coerce')

    # Get the raw data folder for the project
    project_raw_folder = pu.get_project_folder(project_id, RAWDATA_BASE_FOLDER)["path"]
    reporting_week_str_dotted = f"WE {pd.to_datetime(reporting_week, errors='coerce').strftime('%Y.%m.%d')}"
    reporting_week_str = pd.to_datetime(reporting_week, errors='coerce').strftime('%Y-%m-%d')
    weekly_folder = os.path.join(project_raw_folder, reporting_week_str_dotted)

    # get list of required files
    files_required = file_list_compile(project_id, reporting_week_str)

    # check if all files are ready
    if not all_files_ready(files_required, weekly_folder):
        print(f"ERROR: Not all required files are present for project {project_id} - week {reporting_week_str}. Skipping.")
        return False
    
    if project_id == CBV2_PROJECT_ID:
        generate_reports_CB(reporting_week_str, weekly_folder)
    
    elif project_id == EB_PROJECT_ID:
        generate_reports_EB(reporting_week_str, weekly_folder)

    else:
        raise ValueError(f"Unknown project: {project_id}")
    
    
    print(f"Overwritten smr-workflow and smr-rater-label for project {project_id} - week {reporting_week_str}")

    return True



def overwrite_olap_all():
    we_dates = pu.generate_we_dates("2025-01-01")
    for week in we_dates:
        for project_id in [CBV2_PROJECT_ID, EB_PROJECT_ID]:

            reporting_week_str = pd.to_datetime(week, errors="coerce").strftime("%Y-%m-%d")
            olap_folder = os.path.join(OLAP_BASE_FOLDER, project_id, reporting_week_str)

            # verifica se esiste la cartella olap o se è vuota
            if not os.path.exists(olap_folder) or not os.listdir(olap_folder):
                continue

            print(f"Overwriting OLAP for project {project_id} - week {reporting_week_str}")
            overwrite_olap(project_id, reporting_week_str)



#success = overwrite_olap("a01Hs00001ocUa0IAE", "2025-10-03")
#print(success)
#overwrite_olap_all()