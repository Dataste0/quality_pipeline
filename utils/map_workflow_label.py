import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import pipeline_lib.config as cfg
import pipeline_lib.pipeline_utils as pu
from pipeline_lib.sql.query_map_info import query_map_info_run
import pandas as pd
from openpyxl import load_workbook

MARKET_MAP_FILE_PATH = os.path.join(cfg.RAWDATA_ROOT_PATH, "project_info_market_map.xlsx")
LABEL_MAP_FILE_PATH = os.path.join(cfg.RAWDATA_ROOT_PATH, "project_info_label_map.xlsx")

PROJECT_MASTERFILE = cfg.PROJECT_INFO_FILE_PATH

# --- Setup Project List
project_list_df = pu.load_project_info(PROJECT_MASTERFILE, active_only=False)




def append_to_excel(filepath, new_rows_df):
    book = load_workbook(filepath)

    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]

    startrow = sheet.max_row

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        new_rows_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)




def generate_map_info_markets():
    report_df = query_map_info_run("market")
    merged_df = report_df.merge(project_list_df[['project_id', 'project_name']], on='project_id', how='left')

    current_market_info_df = pu.load_df_from_filepath(MARKET_MAP_FILE_PATH)
    
    expected_columns = {'Project Name', 'Mercury ID', 'Routing/Queue'}
    if expected_columns.issubset(current_market_info_df.columns):
        current_df = current_market_info_df.rename(columns={
            "Project Name": "project_name",
            "Mercury ID": "project_id",
            "Routing/Queue": "workflow"
        })
        rename_back = {
            "project_name": "Project Name",
            "project_id": "Mercury ID",
            "workflow": "Routing/Queue"
        }
    else:
        current_df = current_market_info_df.copy()
        rename_back = {}
    
    joined_cols = ['project_id', 'workflow']
    new_entries_df = merged_df.merge(
        current_df[joined_cols].drop_duplicates(),
        on=joined_cols,
        how='left',
        indicator=True
    ).query('_merge == "left_only"').drop(columns=['_merge'])

    if rename_back:
        new_entries_df = new_entries_df.rename(columns=rename_back)
    
    for col in current_market_info_df.columns:
        if col not in new_entries_df.columns:
            new_entries_df[col] = ""

    new_entries_df = new_entries_df[current_market_info_df.columns]

    print(f"Markets - New Entries {len(new_entries_df)}:")
    print("-----------------------")
    print(new_entries_df)

    # Add to excel file: MARKET_MAP_FILE_PATH
    append_to_excel(MARKET_MAP_FILE_PATH, new_entries_df)

    #print(f"Map Markets: added {len(new_entries_df)} new rows.")


def generate_map_info_labels():
    report_df = query_map_info_run("label")
    merged_df = report_df.merge(project_list_df[['project_id', 'project_name']], on='project_id', how='left')

    current_label_info_df = pu.load_df_from_filepath(LABEL_MAP_FILE_PATH)
    
    expected_columns = {'Project Name', 'Mercury ID', 'Original Label'}
    if expected_columns.issubset(current_label_info_df.columns):
        current_df = current_label_info_df.rename(columns={
            "Project Name": "project_name",
            "Mercury ID": "project_id",
            "Original Label": "label"
        })
        rename_back = {
            "project_name": "Project Name",
            "project_id": "Mercury ID",
            "label": "Original Label"
        }
    else:
        current_df = current_label_info_df.copy()
        rename_back = {}
    
    joined_cols = ['project_id', 'label']
    new_entries_df = merged_df.merge(
        current_df[joined_cols].drop_duplicates(),
        on=joined_cols,
        how='left',
        indicator=True
    ).query('_merge == "left_only"').drop(columns=['_merge'])

    if rename_back:
        new_entries_df = new_entries_df.rename(columns=rename_back)
    
    for col in current_label_info_df.columns:
        if col not in new_entries_df.columns:
            new_entries_df[col] = ""

    new_entries_df = new_entries_df[current_label_info_df.columns]

    print(f"Labels - New Entries {len(new_entries_df)}:")
    print("-----------------------")
    print(new_entries_df)

    append_to_excel(LABEL_MAP_FILE_PATH, new_entries_df)

generate_map_info_markets()
#generate_map_info_labels()